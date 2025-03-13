import os
import pandas as pd
from codelib.file_management.dynamic_file_pathing import get_root
import openpyxl

root = get_root()

# Excel file:
excel_path = os.path.join(root, "GBI Optimisation", "TEST.xlsx")

# CSV files:
data_folder = os.path.join(root, "Simulation", "Data")
file_directory = {
    "EUNK": "EUNK_iShares_Core_MSCI_Europe_UCITS_ETF_EUR_Acc.csv",
    "SXR8": "SXR8_iShares_Core_S&P_500_ETF_USD_Acc_EUR.csv",
    "IUSN": "IUSN_iShares_MSCI_World_Small_Cap_UCITS_ETF_USD_Acc.csv",
    "JGHY": "JGHY_JPM_Global_High_Yield_Corporate_Bond_Multi-Factor_UCITS_ETF_USD_acc.csv",
    "LYXF": "LYXF_Amundi_Euro_Government_Bond_15plusY_UCITS_ETF_Acc.csv",
    "SXR4": "SXR4_iShares_MSCI_USA_UCITS_ETF_USD_Acc.csv",
    "SYBA": "SYBA_SPDR_Bloomberg_Euro_Aggregate_Bond_UCITS_ETF_Dist.csv",
    "SYBB": "SYBB_SPDR_Bloomberg_Euro_Government_Bond_UCITS_ETF_Dist.csv",
    "SYBC": "SYBC_SPDR_Bloomberg_Euro_Corporate_Bond_UCITS_ETF_Dist.csv"
}

prices = pd.DataFrame()

# Process each CSV file and extract the first price in 2025 from the 'Slutkurs' column.
for name, path in file_directory.items():
    data_path = os.path.join(data_folder, path)
    data = pd.read_csv(data_path, index_col=0, parse_dates=True, dayfirst=True)
    data.index = pd.to_datetime(data.index, format="%d/%m/%Y")

    # Filter for rows in 2024
    data_2024 = data[data.index.year == 2024]

    if not data_2024.empty:
        # Get the last price in 2024
        last_price_2024 = data_2024.iloc[-1]["Slutkurs"]
        prices.loc[data_2024.index[-1], name] = last_price_2024
    else:
        prices[name] = None

# Now update the existing table in the "Price" sheet without replacing it.
wb = openpyxl.load_workbook(excel_path)
ws = wb["Price"]

# Read the header row (assumed to be in the first row)
header = [cell.value for cell in ws[1]]
# Assume the first column is "Year"
has_year_column = header and header[0] == "Year"

# Determine the row corresponding to 2025 in the table.
target_row = None
if has_year_column:
    for row in ws.iter_rows(min_row=2, max_col=1):
        if row[0].value == 2024:
            target_row = row[0].row
            break

# If no row for 2025 exists, append a new row and set the year in the first column.
if target_row is None:
    target_row = ws.max_row + 1
    if has_year_column:
        ws.cell(row=target_row, column=1, value=2024)

# Update the row with the price data.
# This assumes that, apart from the optional "Year" column, the header contains the asset names.
for col_index, col_name in enumerate(header, start=1):
    # Skip the "Year" column if present
    if has_year_column and col_index == 1:
        continue
    if col_name in prices.columns:
        # Extract the scalar value (e.g. using .iloc[0] from the series)
        series_val = prices[col_name]
        if not series_val.empty:
            cell_value = series_val.iloc[0]
        else:
            cell_value = None
        ws.cell(row=target_row, column=col_index, value=cell_value)

# Save the workbook after updating the table.
wb.save(excel_path)
print("Table updated successfully.")
